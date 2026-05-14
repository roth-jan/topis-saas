#!/usr/bin/env python3
"""Februar 2026 Scans-Excel aus SharePoint ziehen + zur TOPIS-CSV bauen."""
import json, urllib.request, urllib.parse, csv
from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime

CRED = json.loads(Path.home().joinpath('.openclaw/workspace/sharepoint_credentials.json').read_text())

def token():
    url = f"https://login.microsoftonline.com/{CRED['tenant_id']}/oauth2/v2.0/token"
    data = urllib.parse.urlencode({'client_id': CRED['client_id'],'client_secret': CRED['client_secret'],
        'scope': 'https://graph.microsoft.com/.default','grant_type': 'client_credentials'}).encode()
    with urllib.request.urlopen(urllib.request.Request(url, data=data)) as r:
        return json.loads(r.read())['access_token']

TOK = token()
def g(url):
    if not url.startswith('http'): url = 'https://graph.microsoft.com/v1.0' + url
    with urllib.request.urlopen(urllib.request.Request(url, headers={'Authorization': f'Bearer {TOK}'})) as r:
        return json.loads(r.read())

site = g('/sites/ntcdin.sharepoint.com:/sites/Logistik-Beratung')
drive_id = next(d for d in g(f"/sites/{site['id']}/drives")['value'] if d['name'] == 'Documents')['id']

path = 'General/Logistikberatung/Andreas Schmid 2026/DatenExAS/Auswertung TW/Scans Februar 26.xlsx'
enc = '/'.join(urllib.parse.quote(seg, safe='') for seg in path.split('/'))
item = g(f"/drives/{drive_id}/root:/{enc}")
print(f"Datei: {item['name']}  ({item.get('size',0)/1024/1024:.1f} MB)")

dst = Path('/tmp/as-feb2026')
dst.mkdir(exist_ok=True)
xlsx = dst / 'scans_feb2026.xlsx'
url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item['id']}/content"
with urllib.request.urlopen(urllib.request.Request(url, headers={'Authorization': f'Bearer {TOK}'})) as r:
    xlsx.write_bytes(r.read())
print(f"OK Download → {xlsx} ({xlsx.stat().st_size/1024/1024:.1f} MB)")

# Excel laden, erste sheet
print("Lade Workbook (kann dauern)...")
wb = load_workbook(xlsx, data_only=True, read_only=True)
print(f"Sheets: {wb.sheetnames}")
ws = wb['Raw']  # Raw-Scans, nicht Strecken-Auswertung

# Header lesen
rows_iter = ws.iter_rows(values_only=True)
header = [str(h).strip().lower() if h else '' for h in next(rows_iter)]
print(f"Header: {header}")

# Spalten-Indizes finden
def col(name):
    for i, h in enumerate(header):
        if name in h: return i
    return -1

col_mp = col('messpunkt')
col_date = col('scandatum')
col_time = col('scanzeit')
col_tour = col('tour')
col_disp = col('dispogebiet')
col_kolli = col('kolli')
col_tkg = col('tkg')
col_auftrag = col('auftragsnr_eindeutig')
col_stellplatz = col('stellplatz')

print(f"Indizes: mp={col_mp}, date={col_date}, time={col_time}, tour={col_tour}, kolli={col_kolli}, tkg={col_tkg}")

MP_NAME = {
    '5':  'SE Entladung FV', '7':  'SA Verladung FV',
    '4':  'Nat. Ausgang',   '4a': 'Sonderausgang',
    '2':  'Nat. Eingang',    '9b': 'NV Eingang',
    '9a': 'Sonderscan',
}

aggr = defaultdict(lambda: {'sendungen': set(), 'colli': 0, 'gewicht': 0.0, 'disp': '', 'stellplatz': ''})

n_total = 0
n_skipped = 0
for row in rows_iter:
    n_total += 1
    if col_mp < 0 or col_mp >= len(row): continue
    mp_raw = row[col_mp]
    if mp_raw is None: n_skipped += 1; continue
    mp = str(mp_raw).strip()
    if not mp: n_skipped += 1; continue
    d = row[col_date]
    if d is None: n_skipped += 1; continue
    if isinstance(d, datetime):
        datum = d.strftime('%Y-%m-%d')
    else:
        datum = str(d)[:10]
    t = row[col_time] if col_time >= 0 else None
    if isinstance(t, datetime):
        zeit = t.strftime('%H:%M')
    else:
        zeit = '00:00'
    h, m = zeit.split(':')
    halbstunde = f"{h}:{'30' if int(m) >= 30 else '00'}"
    tour = str(row[col_tour] or '').strip() if col_tour >= 0 else ''
    disp = str(row[col_disp] or '').strip() if col_disp >= 0 else ''
    stellplatz = str(row[col_stellplatz] or '').strip() if col_stellplatz >= 0 else ''
    try:
        kolli = int(row[col_kolli] or 1) if col_kolli >= 0 else 1
    except (ValueError, TypeError):
        kolli = 1
    try:
        gewicht = float(row[col_tkg] or 0) if col_tkg >= 0 else 0
    except (ValueError, TypeError):
        gewicht = 0
    auftrag = str(row[col_auftrag] or '').strip() if col_auftrag >= 0 else ''

    key = (datum, halbstunde, mp, tour)
    a = aggr[key]
    if auftrag: a['sendungen'].add(auftrag)
    a['colli'] += kolli
    a['gewicht'] += gewicht
    if disp: a['disp'] = disp
    if stellplatz: a['stellplatz'] = stellplatz

print(f"Verarbeitet: {n_total} (übersprungen: {n_skipped}), Aggregate: {len(aggr)}")

OUT = '/Users/clawdette/projects/topis-saas/public/demo-data/as-feb2026-scans.csv'
import os
os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f, delimiter=';')
    w.writerow(['scandatum','scanzeit','messpunkt','messpunktname','tour','dispogebiet','sendungen','colli','gewicht','ladezeit'])
    for (datum, zeit, mp, tour), a in sorted(aggr.items()):
        mp_norm = mp if mp.startswith('MP') else f'MP{mp}'
        mp_clean = mp_norm.replace('MP', '')
        mp_name = MP_NAME.get(mp_clean, f'MP{mp_clean}')
        w.writerow([datum, zeit, mp_norm, mp_name, tour, a['disp'],
                    len(a['sendungen']), a['colli'], round(a['gewicht'], 1), ''])

print(f"→ {OUT}  ({os.path.getsize(OUT)/1024:.1f} KB)")
